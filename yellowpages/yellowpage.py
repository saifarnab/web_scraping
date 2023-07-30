import os
import time
from urllib.parse import unquote

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By


# install dependencies
# subprocess.check_call(['pip', 'install', 'pillow'])
# subprocess.check_call(['pip', 'install', 'requests'])
# subprocess.check_call(['pip', 'install', 'requests'])
# subprocess.check_call(['pip', 'install', 'openpyxl'])
# subprocess.check_call(['pip', 'install', 'selenium'])
# subprocess.check_call(['pip', 'install', 'fake_useragent'])


def config_driver() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--start-miniimized")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    # chrome_options.add_argument(f'user-agent={UserAgent().random}')
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def append_to_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def create_excel_with_header(filename):
    if os.path.exists(filename) is True:
        print(f'{filename} already exist')
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Industry', 'CompanyName', 'Phone', 'Website'])
    workbook.save(filename)
    print(f'{filename.title()} created')


def check_data_exists(filename: str, company_name: str, phone_number: str) -> bool:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[1] is None:
            return False
        if company_name.strip() == row[1].strip() and phone_number == row[2]:
            return True  # Data exists in the sheet
    return False  # Data does not exist in the sheet


def scrapper():
    print('Script starts running... ')
    filename = 'yellow_pre.xlsx'
    create_excel_with_header(filename)
    driver = config_driver()

    industries = []

    for url in industries:
        driver.get(url.replace('page_var', '1'))
        max_page = driver.find_element(By.XPATH, '//span[@class="pageCount"]').text.split('/')[1].strip()
        for i in range(1, int(max_page) + 1):
            driver.get(url.replace('page_var', str(i)))
            items = driver.find_elements(By.XPATH, '//div[@class="listing_right_section"]')
            for item in items:
                industry_name_xpath = "//h1[contains(@class,'page__container-title')]//strong[1]"
                company_name_xpath = './/a[@class="listing__name--link listing__link js ListingName"]'
                phone_number_xpath = ".//a[@title='Get the Phone Number']"
                website_xpath = ".//li[@class='mlr__item mlr__item- -website ']//a"

                try:
                    industry_name = driver.find_element(By.XPATH, industry_name_xpath).text.strip()
                except:
                    industry_name = ''

                try:
                    company_name = item.find_element(By.XPATH, company_name_xpath).text.strip()
                except:
                    company_name = ''
                try:
                    phone_number = item.find_element(By.XPATH, phone_number_xpath).get_attribute('data-phone').strip()
                except:
                    phone_number = ''
                try:
                    website = item.find_element(By.XPATH, website_xpath)
                except:
                    website = ''

                if check_data_exists(filename, company_name, phone_number) is False:
                    append_to_excel(filename, [[industry_name, company_name, phone_number, website]])
                    print(f'{industry_name} -> {i} -> {company_name} -> inserted.')

            print('---------------------------------------------------')

    print('Execution done!')


if __name__ == '__main__':
    scrapper()
