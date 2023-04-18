import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('data.xlsx')
ws = wb['Sheet1']
source = ws['D18':'G20']
table2 = ws['Q18':'T20']
table3 = ws['AB18':'AE20']

data = []

counter = 0
sl = 1
pre1, pre2 = None, "Lettuce lettings, 17 Raglan Street, Coventry, CV1 5QF"


data.append(str(sl))
for no, name, address in zip(ws['A'], ws['F'], ws['G']):
    if counter > 1:
        if name.value == pre1 and address.value == pre2:
            data.append(" ")
            continue
        sl += 1
        data.append(str(sl))

    counter += 1
    pre1 = name.value
    pre2 = address.value

columns = ['no']
df = pd.DataFrame(data, columns=columns)
df.to_excel(f"data_re_gen.xlsx", index=False)