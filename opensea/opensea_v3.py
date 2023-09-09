import os
import subprocess
import time
from io import BytesIO

import openpyxl
import pandas as pd
import requests
from PIL import Image
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

# install dependencies
# subprocess.check_call(['pip', 'install', 'pillow'])
# subprocess.check_call(['pip', 'install', 'requests'])
# subprocess.check_call(['pip', 'install', 'openpyxl'])
# subprocess.check_call(['pip', 'install', 'selenium'])
# subprocess.check_call(['pip', 'install', 'fake_useragent'])


def config_driver() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument(f'user-agent={UserAgent().random}')
    # chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def config_driver_without_ua() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument(f'user-agent={UserAgent().random}')
    # chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def create_directory_if_not_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def save_nft(src: str, name: str):
    if src.split('.')[-1] == 'avif':
        reloader = 1
        while True:
            if reloader >= 5:
                return None
            response = requests.get(src)
            if response.status_code == 200:
                im = Image.open(BytesIO(response.content))
                rgb_im = im.convert('RGB')
                filename = f'assets/{name}.jpg'
                rgb_im.save(filename)
                return filename
            reloader += 1

    else:
        if '.png' in src:
            file_name = f'assets/{name}.png'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
            return file_name
        elif '.mp4' in src:
            file_name = f'assets/{name}.mp4'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)

            return file_name
        elif '.webp' in src:
            file_name = f'assets/{name}.webp'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
            return file_name
        elif '.jpg' in src:
            file_name = f'assets/{name}.jpg'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
            return file_name
        elif '.gif' in src:
            file_name = f'assets/{name}.gif'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
            return file_name
        elif '.jpg' in src:
            file_name = f'assets/{name}.jpg'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
            return file_name
    return None


def append_to_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def create_excel_with_header(filename):
    if os.path.exists(filename) is True:
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Link', 'Name', 'Image Name'])
    workbook.save(filename)


def get_last_inserted_nft(filename):
    try:
        df = pd.read_excel(filename, skiprows=1)
        df = df.dropna(how='all')
        last_non_blank_row = df.tail(1)
        return str(last_non_blank_row.values[0][0]).split('/')[-1]
    except:
        return None


def check_data_exists(filename: str, name: str) -> bool:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[1] is None:
            return False
        if name.strip() == row[1].strip():
            return True  # Data exists in the sheet
    return False  # Data does not exist in the sheet


def is_access_denied(driver) -> bool:
    try:
        driver.find_element(By.XPATH, '//div[@class="cf-error-title"]')
        return True
    except Exception as e:
        return False


def _take_input_url() -> str:
    return input('Enter oldest detail url: ').strip()


def _generate_base_url(url: str) -> (str, int):
    base_url = url.rsplit('/', 1)[0]
    oldest = url.split('/')[-1]
    return base_url, int(oldest)


def scrapper():
    print('=============================================================')
    url = _take_input_url()
    base_url, oldest = _generate_base_url(url)
    init, reloader, name, filename = 0, 1, '', ''
    print('Starting script ... ')
    while True:

        if reloader > 1:
            print('Reloading ... ')
        if reloader >= 5:
            break

        if init == 0:
            try:
                driver = config_driver()
                driver.get(url)
                collection_name = driver.find_element(By.XPATH, '//div[@class="item--collection-detail"]//div//div//a').text
                collection_name = collection_name.split('/')[-1].replace('-', ' ').upper()
                print(collection_name)
                WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.CLASS_NAME, 'item--header')))
            except Exception as e:
                try:
                    driver = config_driver_without_ua()
                    driver.get(url)
                    collection_name = driver.find_element(By.XPATH,
                                                          '//div[@class="item--collection-detail"]//div//div//a').text
                    collection_name = collection_name.split('/')[-1].replace('-', ' ').upper()
                    print(collection_name)
                    WebDriverWait(driver, 5).until(
                        EC.visibility_of_element_located((By.CLASS_NAME, 'item--header')))
                except Exception as ex:
                    reloader += 1
                    continue
            # collection_name = driver.find_element(By.XPATH, '//a[@class="sc-1f719d57-0 eiItIQ CollectionLink--link"]//span//div').text
            create_directory_if_not_exists(f'assets/{collection_name}')
            create_directory_if_not_exists('files')
            filename = f'files/{collection_name}.xlsx'
            create_excel_with_header(filename)
            if get_last_inserted_nft(filename) is not None:
                oldest = int(get_last_inserted_nft(filename)) + 1
            init = 1

        else:
            try:
                driver = config_driver()
                collection_name = driver.find_element(By.XPATH,
                                                      '//div[@class="item--collection-detail"]//div//div//a').text
                collection_name = collection_name.split('/')[-1].replace('-', ' ').upper()
                print(collection_name)
                driver.get(f'{base_url}/{oldest}')
                WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.XPATH, '//img[@class="Image--image"]')))
                src = driver.find_element(By.XPATH, '//img[@class="Image--image"]').get_attribute('src')

            except Exception as e:
                driver = config_driver_without_ua()
                driver.get(f'{base_url}/{oldest}')
                try:
                    WebDriverWait(driver, 5).until(
                        EC.visibility_of_element_located((By.XPATH, '//img[@class="Image--image"]')))
                    src = driver.find_element(By.XPATH, '//img[@class="Image--image"]').get_attribute('src')

                except Exception as ex:
                    driver = config_driver()
                    driver.get(f'{base_url}/{oldest}')
                    try:
                        time.sleep(1)
                        src = driver.find_element(By.TAG_NAME, "source").get_attribute('src')
                    except Exception as e:
                        driver = config_driver_without_ua()
                        driver.get(f'{base_url}/{oldest}')
                        try:
                            time.sleep(1)
                            src = driver.find_element(By.TAG_NAME, "source").get_attribute('src')
                        except Exception as ex:
                            reloader += 1
                            continue

            link = driver.current_url
            collection_name = driver.find_element(By.XPATH,
                                                  '//div[@class="item--collection-detail"]//div//div//a').text
            collection_name = collection_name.split('/')[-1].replace('-', ' ').upper()
            print(collection_name)
            temp_name = driver.find_element(By.XPATH, "//span[@aria-expanded='false']//div").text + f'#{oldest}'

            if check_data_exists(filename, temp_name) is False:
                asset_name = save_nft(src, f'{collection_name}/{temp_name}')
                if asset_name is not None:
                    append_to_excel(filename, [[link, temp_name, asset_name.split('/')[-1]]])
                    print(f'--> {temp_name} inserted')
            else:
                print(f'--> {temp_name} is already available.')

            oldest += 1
            reloader = 1

    print('Execution done!')


if __name__ == '__main__':
    scrapper()
