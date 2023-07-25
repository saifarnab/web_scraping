import os
import subprocess
import time
from io import BytesIO
import urllib

import openpyxl
import pandas as pd
import requests
from PIL import Image
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


# install dependencies
# subprocess.check_call(['pip', 'install', 'pillow'])
# subprocess.check_call(['pip', 'install', 'requests'])
# subprocess.check_call(['pip', 'install', 'openpyxl'])
# subprocess.check_call(['pip', 'install', 'selenium'])
# subprocess.check_call(['pip', 'install', 'fake_useragent'])


def config_driver() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument(f'user-agent={UserAgent().random}')
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def config_driver_without_ua() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def create_directory_if_not_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Directory '{directory}' created.")
    else:
        print(f"Directory '{directory}' already exist.")


def save_nft(src: str, name: str):
    if src.split('.')[-1] == 'avif':
        reloader = 1
        while True:
            if reloader >= 5:
                return False
            response = requests.get(src)
            if response.status_code == 200:
                im = Image.open(BytesIO(response.content))
                rgb_im = im.convert('RGB')
                rgb_im.save(f'images/{name}.jpg')
                return True
            reloader += 1

    else:
        if '.png' in src:
            file_name = f'images/{name}.png'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
        elif '.mp4' in src:
            file_name = f'images/{name}.mp4'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)
        elif '.webp' in src:
            file_name = f'images/{name}.webp'
            resp = requests.get(src)
            with open(file_name, "wb") as f:
                f.write(resp.content)


def append_to_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def create_excel_with_header(filename):
    if os.path.exists(filename) is True:
        print(f'{filename} already exist')
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Link', 'Name', 'Image Name'])
    workbook.save(filename)
    print(f'{filename.title()} created')


def get_last_inserted_nft(filename):
    try:
        df = pd.read_excel(filename, skiprows=1)
        df = df.dropna(how='all')
        last_non_blank_row = df.tail(1)
        return str(last_non_blank_row.values[0][0]).split('/')[-1]
    except:
        return None


def check_data_exists(filename: str, name: str) -> bool:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[1] is None:
            return False
        if name.strip() == row[1].strip():
            return True  # Data exists in the sheet
    return False  # Data does not exist in the sheet


def is_access_denied(driver) -> bool:
    try:
        driver.find_element(By.XPATH, '//div[@class="cf-error-title"]')
        return True
    except Exception as e:
        return False


def _take_input_url() -> str:
    return input('Enter oldest detail url: ')


def _generate_base_url(url: str) -> (str, int):
    base_url = url.rsplit('/', 1)[0]
    oldest = url.split('/')[-1]
    return base_url, int(oldest)


def scrapper():
    print('=============================================================')
    url = _take_input_url()
    base_url, oldest = _generate_base_url(url)
    init, reloader, name, filename = 0, 1, '', ''
    print('Starting data extraction ... ')
    while True:
        print(oldest)
        if init == 0:
            try:
                driver = config_driver()
                driver.get(url)
                try:
                    WebDriverWait(driver, 3).until(
                        EC.visibility_of_element_located((By.CLASS_NAME, 'item--header')))
                except Exception as e:
                    reloader += 1
                    continue

                name = driver.find_element(By.TAG_NAME, 'h1').text.split('#')[0].strip()
                create_directory_if_not_exists(f'images/{name}')
                create_directory_if_not_exists('files')
                filename = f'files/{name}.xlsx'
                create_excel_with_header(filename)
                if get_last_inserted_nft(filename) is not None:
                    oldest = int(get_last_inserted_nft(filename)) + 1

            except:
                try:
                    driver = config_driver_without_ua()
                    driver.get(url)
                    name = driver.find_element(By.TAG_NAME, 'h1').text.split('#')[0].strip()
                    create_directory_if_not_exists(f'images/{name}')
                    create_directory_if_not_exists('files')
                    filename = f'files/{name}.xlsx'
                    if get_last_inserted_nft(filename) is not None:
                        oldest = int(get_last_inserted_nft(filename)) + 1
                except:
                    reloader += 1
                    continue
            init = 1

        else:
            driver = config_driver()
            driver.get(f'{base_url}/{oldest}')

            if reloader >= 5:
                break

            try:
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//img[@class="Image--image"]')))
            except Exception as e:
                # print('Retrying without fake user-agent chrome driver')
                driver = config_driver_without_ua()
                driver.get(f'{base_url}/{oldest}')
                try:
                    WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.XPATH, '//img[@class="Image--image"]')))
                except Exception as ex:
                    # print('Retrying with fake user-agent chrome driver')
                    reloader += 1
                    continue

            link = driver.current_url
            temp_name = driver.find_element(By.XPATH, '//h1[@class="sc-29427738-0 ivioUu item--title"]').text
            src = driver.find_element(By.XPATH, '//img[@class="Image--image"]').get_attribute('src')
            print(temp_name, src)
            if check_data_exists(filename, temp_name) is False:
                if save_nft(src, f'{name}/{temp_name}') is True:
                    append_to_excel(filename, [[link, temp_name, f'{temp_name}.jpg']])
                    print(f'--> {temp_name} inserted')
            else:
                print(f'--> {temp_name} is already available.')

            oldest += 1
            reloader = 1

    print('Execution done!')


if __name__ == '__main__':
    scrapper()
