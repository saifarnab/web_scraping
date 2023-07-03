import os
import subprocess
import time
from io import BytesIO

import openpyxl
import requests
from PIL import Image
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys


# install dependencies
subprocess.check_call(['pip', 'install', 'pillow'])
subprocess.check_call(['pip', 'install', 'requests'])
subprocess.check_call(['pip', 'install', 'requests'])
subprocess.check_call(['pip', 'install', 'openpyxl'])
subprocess.check_call(['pip', 'install', 'selenium'])
subprocess.check_call(['pip', 'install', 'fake_useragent'])


def config_driver() -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--start-miniimized")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument(f'user-agent={UserAgent().random}')
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    return driver


def scroll_down_page(driver):
    html = driver.find_element(By.TAG_NAME, 'html')
    for i in range(100):
        html.send_keys(Keys.ARROW_DOWN)


def create_directory_if_not_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Directory '{directory}' created.")
    else:
        print(f"Directory '{directory}' already exist.")


def save_as_jpg(src: str, name: str):
    response = requests.get(src)
    im = Image.open(BytesIO(response.content))
    rgb_im = im.convert('RGB')
    rgb_im.save(f'images/{name}.jpg')


def append_to_excel(filename, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def create_excel_with_header(filename):
    if os.path.exists(filename) is True:
        print(f'{filename} already exist')
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['No.', 'Link', 'Name', 'Image Name'])
    workbook.save(filename)
    print(f'{filename.title()} created')


def get_last_row(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    last_row = sheet.max_row
    return last_row


def check_data_exists(filename: str, name: str) -> bool:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if name.strip() == row[2].strip():
            return True  # Data exists in the sheet
    return False  # Data does not exist in the sheet


def is_access_denied(driver) -> bool:
    try:
        driver.find_element(By.XPATH, '//div[@class="cf-error-title"]')
        return True
    except Exception as e:
        return False


def scrapper():
    print('-------------------------------------------------------------')
    print('-------------------------------------------------------------')
    print('Script starts running... ')
    filename = 'opensea.xlsx'
    create_directory_if_not_exists('images')
    create_excel_with_header(filename)
    azuki = 0
    reloader = 0
    no = get_last_row(filename)
    print('Starting data extraction ... ')
    while True:
        if azuki >= 10000:
            break
        print(f'Azuki = {azuki}')
        driver = config_driver()
        driver.get(
            f'https://opensea.io/collection/azuki?search[query]=Azuki%20%23{azuki}')
        time.sleep(5)
        if is_access_denied(driver) is True:
            continue
        scroll_down_page(driver)
        articles = driver.find_elements(By.XPATH,
                                        '//div[@class="sc-29427738-0 sc-7ac0e573-1 cKdnBO iCfZHJ Asset--loaded"]')

        if len(articles) == 0 and reloader <= 5:
            reloader += 1
            continue
        for article in articles:
            link = driver.find_element(By.XPATH, './/article/a').get_attribute('href')
            img = article.find_element(By.XPATH, './/img')
            src = img.get_attribute('src')
            name = img.get_attribute('alt')
            if 'https' in src:
                if check_data_exists(filename, name) is False:
                    save_as_jpg(src, name)
                    append_to_excel(filename, [[no, link, name, f'{name}.jpg']])
                    print(f'--> {no}. {name} inserted')
                    no += 1
                else:
                    print(f'--> {name} is already available.')

        azuki += 1
        reloader = 0

    print('Execution done!')


if __name__ == '__main__':
    scrapper()
