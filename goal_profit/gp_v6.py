import csv
import logging
import os
import random
import subprocess
import time
from datetime import date

import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import xlwings as xw


# subprocess.check_call(['pip', 'install', 'selenium'])
# subprocess.check_call(['pip', 'install', 'openpyxl'])
# subprocess.check_call(['pip', 'install', 'pandas'])

# excel db path
GAMEDB = 'scrap.xlsx'

# log format
logging.basicConfig(
    format='%(asctime)s %(levelname)s %(message)s',
    datefmt='%H:%M:%S',
    level=logging.INFO)


def config_driver(maximize_window: bool) -> webdriver.Chrome:
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("lang=en-GB")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--proxy-server='direct://'")
    chrome_options.add_argument("--proxy-bypass-list=*")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=chrome_options)
    if maximize_window is True:
        driver.maximize_window()
    return driver


def check_excel_file() -> bool:
    if os.path.exists(GAMEDB):
        return True
    return False


def create_csv(filename: str):
    file_exists = os.path.isfile(filename)
    headers = ['Provider', 'SelectionName', 'MarketName', 'EventName', 'MarketType', 'BetType']
    if not file_exists:
        with open(filename, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow(headers)


def get_last_data_from_csv(filename: str) -> list:
    try:
        df = pd.read_csv(filename)
        return df.iloc[-1].tolist()
    except:
        return []


def save_csv_data(filename: str, row: list):
    df = pd.read_csv(filename)
    new_row_df = pd.DataFrame([row], columns=df.columns)
    df = pd.concat([df, new_row_df], ignore_index=True)
    df.to_csv(filename, index=False)


def get_first_blank_row():
    workbook = openpyxl.load_workbook(GAMEDB)
    sheet = workbook['Scrape']
    first_blank_row = sheet.max_row + 1  # Default to the first row if no data in the sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if all(cell.value is None for cell in row):
            first_blank_row = row[0].row
            break

    return first_blank_row


# Function to find the first blank row in a chunk
def find_first_blank_row_in_chunk(chunk):
    for row_num, row in enumerate(chunk, start=1):
        print(row_num)
        if all(cell is None for cell in row):
            return row_num
    return None


def get_first_blank_row3():
    df = pd.read_excel(GAMEDB, engine="openpyxl")
    return df.shape[0] + 1


def get_first_blank_row2():
    print('hit 1')
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(GAMEDB)
    sheet = workbook['Scrape']
    chunk_size = 1000
    first_blank_row = 0
    print('hit 2')
    # Iterate through rows in chunks
    for row_start in range(1, sheet.max_row + 1, chunk_size):
        row_end = min(row_start + chunk_size - 1, sheet.max_row)
        row_chunk = sheet.iter_rows(min_row=row_start, max_row=row_end, values_only=True)

        blank_row_in_chunk = find_first_blank_row_in_chunk(row_chunk)
        if blank_row_in_chunk is not None:
            first_blank_row = row_start + blank_row_in_chunk - 1
            break

    return first_blank_row


def insert_or_update_row(first_blank_row: int, new_data: list) -> bool:
    df = pd.read_excel(GAMEDB, sheet_name='Scrape', skiprows=1)
    # is_row_available = ((df['GameDate'] == new_data[0]) & (df['Game'] == new_data[1])).any()
    # if is_row_available:
    #     return False
    workbook = openpyxl.load_workbook(GAMEDB)
    sheet = workbook['Scrape']
    start_row = first_blank_row  # Append after the last row in the sheet
    start_col = 1  # Start from the first column (A)
    for col_idx, value in enumerate(new_data, start=start_col):
        sheet.cell(row=start_row, column=col_idx, value=value)
    sheet.append(new_data)
    workbook.save(GAMEDB)
    workbook.close()
    return True


def insert_or_update_row2(first_blank_row: int, new_data: list) -> bool:
    workbook = openpyxl.load_workbook(GAMEDB)
    sheet = workbook["Scrape"]
    data_exists = any(all(cell.value == value for cell, value in zip(row, new_data)) for row in sheet.iter_rows())

    if not data_exists:
        for col_num, value in enumerate(new_data, start=1):
            sheet.cell(row=first_blank_row, column=col_num, value=value)
        workbook.save(GAMEDB)
        return True
    else:
        return False


def insert_or_update_row3(first_blank_row: int, new_data: list) -> bool:
    wb = xw.Book(GAMEDB)
    sheet = wb.sheets["Scrape"]

    data_exists = any(all(cell.value == value for cell, value in zip(row, new_data)) for row in
                      sheet.range(f"A1:A{first_blank_row}").rows)

    if not data_exists:
        row_range = sheet.range(f"A{first_blank_row}:O{first_blank_row}")
        for cell, value in zip(row_range, new_data):
            cell.value = value
        wb.save()
        wb.close()
        return True
    else:
        wb.close()
        return False


def save_to_csv():
    # for HT 0-0
    try:
        df = pd.read_excel(GAMEDB, sheet_name='HT 0-0', skiprows=1)
        df = df[['Provider', 'SelectionName', 'MarketName', 'EventName', 'MarketType', 'BetType']]
        last_non_blank_row_index = df.last_valid_index()
        last_non_blank_row_data = df.iloc[last_non_blank_row_index]
        last_excel_row = last_non_blank_row_data.tolist()
        create_csv('HT 0-0.csv')
        last_csv_row = get_last_data_from_csv('HT 0-0.csv')
        if last_excel_row != last_csv_row:
            save_csv_data('HT 0-0.csv', last_excel_row)
    except:
        pass

    # for HT 0-1
    try:
        df = pd.read_excel(GAMEDB, sheet_name='HT 0-1', skiprows=1)
        df = df[['Provider', 'SelectionName', 'MarketName', 'EventName', 'MarketType', 'BetType']]
        last_non_blank_row_index = df.last_valid_index()
        last_non_blank_row_data = df.iloc[last_non_blank_row_index]
        last_excel_row = last_non_blank_row_data.tolist()
        create_csv('HT 0-1.csv')
        last_csv_row = get_last_data_from_csv('HT 0-1.csv')
        if last_excel_row != last_csv_row:
            save_csv_data('HT 0-1.csv', last_excel_row)
    except:
        pass

    # for HT 1-0
    try:
        df = pd.read_excel(GAMEDB, sheet_name='HT 1-0', skiprows=1)
        df = df[['Provider', 'SelectionName', 'MarketName', 'EventName', 'MarketType', 'BetType']]
        last_non_blank_row_index = df.last_valid_index()
        last_non_blank_row_data = df.iloc[last_non_blank_row_index]
        last_excel_row = last_non_blank_row_data.tolist()
        create_csv('HT 1-0.csv')
        last_csv_row = get_last_data_from_csv('HT 1-0.csv')
        if last_excel_row != last_csv_row:
            save_csv_data('HT 1-0.csv', last_excel_row)
    except:
        pass

    # for HT 1-1
    try:
        df = pd.read_excel(GAMEDB, sheet_name='HT 1-1', skiprows=1)
        df = df[['Provider', 'SelectionName', 'MarketName', 'EventName', 'MarketType', 'BetType']]
        last_non_blank_row_index = df.last_valid_index()
        last_non_blank_row_data = df.iloc[last_non_blank_row_index]
        last_excel_row = last_non_blank_row_data.tolist()
        create_csv('HT 1-1.csv')
        last_csv_row = get_last_data_from_csv('HT 1-1.csv')
        if last_excel_row != last_csv_row:
            save_csv_data('HT 1-1.csv', last_excel_row)
    except:
        pass


def scanner():
    # define 'driver' variable
    if check_excel_file() is False:
        return

    first_blank_row = get_first_blank_row()
    driver = config_driver(True)

    # iterate to extract game data
    while True:

        try:
            driver.get('https://live.goalprofits.com/')
            # print(f'{pathlib.Path(__file__).parent.resolve()}\gamedb.gamedb;')

            if 'login' in driver.current_url:
                # login to the site
                username_input_button = driver.find_element(By.XPATH, '//input[@id="wlm_form_field_log"]')
                username_input_button.clear()
                username_input_button.send_keys("dennis.peibst@googlemail.com")
                pass_input_button = driver.find_element(By.XPATH, '//input[@id="wlm_form_field_pwd"]')
                pass_input_button.clear()
                pass_input_button.send_keys("Freelancer1")
                driver.find_element(By.XPATH, '//input[@id="wlm_form_field_wp-submit"]').click()
                logging.info("login success")
                time.sleep(2)
                # get to the live game section
                driver.get('https://live.goalprofits.com/')

            elif 'live' not in driver.current_url:
                logging.info('login or live not found')
                continue

            live_games_button = driver.find_element(By.XPATH, '//a[@href="#live"]')
            driver.execute_script("arguments[0].click();", live_games_button)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", live_games_button)
            time.sleep(1)

            # find available live games
            table_elements = driver.find_elements(By.XPATH, '//div[@id="html2"]//table')

            # wait till 60s if no live games are available
            if len(table_elements) == 0:
                logging.info('--> no live games are playing, waiting 60s for next try.')
                time.sleep(60)
                continue

            # iterate each tr
            for table_ind in range(len(table_elements)):
                tr_elements = table_elements[table_ind].find_elements(By.XPATH, './/tbody//tr')
                tr_ind = 0

                # iterate each td
                for ind in range(int(len(tr_elements) / 2)):
                    td_elements = tr_elements[tr_ind].find_elements(By.XPATH, ".//td")
                    current_date = date.today().strftime('%d/%m/%y')
                    try:
                        game_time = td_elements[0].text.strip()
                        logging.info(f'game_time --> {game_time}')
                    except Exception as exx:
                        game_time = 'N/A'

                    # if game time is not HT then return
                    # if game_time != 'HT':
                    #     tr_ind += 2
                    #     logging.info('Waiting to detect HT')
                    #     continue

                    # extract required data from 1st tr
                    team1 = td_elements[1].text.index('(')
                    team1 = td_elements[1].text[:team1].strip()
                    logging.info(f'team1 --> {team1}')

                    try:
                        ht_score = td_elements[2].text.split('\n')[0]
                        logging.info(f'ht_score --> {ht_score}')
                    except Exception as exx:
                        ht_score = 'N/A'

                    try:
                        v = td_elements[19].find_element(By.XPATH, './/a/img[@data-toggle="tooltip"]')
                        tooltip = v.get_attribute('data-original-title').replace('\n', '').replace(
                            "<div class='tooltip-prices'>", '').replace('<br>', '').replace('</div>', '').replace(
                            '<div>', '')
                        tooltip = "".join(tooltip.split())
                        ht_home, ht_draw, ht_away = tooltip[2:6], tooltip[8:12], tooltip[14:18]
                    except Exception as e:
                        ht_home, ht_draw, ht_away = 'N/A', 'N/A', 'N/A'

                    try:
                        home_on = td_elements[3].text
                        logging.info(f'home_on --> {home_on}')
                    except Exception as exx:
                        home_on = 'N/A'

                    try:
                        home_off = td_elements[4].text
                        logging.info(f'home_off --> {home_off}')
                    except Exception as exx:
                        home_off = 'N/A'

                    try:
                        home_da = td_elements[9].text
                        logging.info(f'home_da --> {home_da}')
                    except Exception as exx:
                        home_da = 'N/A'

                    logging.info('extracted required data from 1st tr')

                    # extract required data from 2nd tr
                    td_elements2 = tr_elements[tr_ind + 1].find_elements(By.XPATH, ".//td")

                    team2 = td_elements2[0].text.index('(')
                    team2 = td_elements2[0].text[:team2].strip()
                    # logging.info(f'team2 --> {team2}')

                    try:
                        away_on = td_elements2[1].text
                        logging.info(f'away_on --> {away_on}')
                    except Exception as exx:
                        away_on = 'N/A'

                    try:
                        away_off = td_elements2[2].text
                        logging.info(f'away_off --> {away_off}')
                    except Exception as exx:
                        away_off = 'N/A'

                    try:
                        away_da = td_elements2[7].text
                        logging.info(f'away_da --> {away_da}')
                    except Exception as exx:
                        away_da = 'N/A'

                    # logging.info('extracted required data from 2nd tr')

                    game = team1 + ' v ' + team2

                    # extract data from i button which will open a modal
                    # td_elements[2].find_element(By.XPATH, './/a[2]').click()
                    driver.find_element(By.XPATH,
                                        "//img[@src='https://live.goalprofits.com/_layouts/images/info.png']").click()

                    # logging.info(f'i button 1st click done')
                    time.sleep(1)
                    driver.find_element(By.XPATH, '//a[@id="pregame-tab"]').click()
                    # logging.info(f'i button 2nd click done')
                    time.sleep(1)
                    try:
                        home = driver.find_elements(By.XPATH,
                                                    '//table[@class="table borderless table-striped score-table"]//span[@class="d-block font13 text-nowrap"]')[
                            0].text.split(':')[1].strip()
                        logging.info(f'home --> {home}')
                    except Exception as exx:
                        home = 'N/A'

                    # close the modal opened by I button
                    close_buttons = driver.find_elements(By.XPATH, '//button[@class="close"]')
                    for close_button in close_buttons:
                        try:
                            close_button.click()
                            # logging.info('clieck close button to close the modal')
                        except Exception as e:
                            continue

                    time.sleep(1)

                    # insert data to gamedb
                    insert = insert_or_update_row(first_blank_row,
                                                  [current_date, game, home, '', '', ht_home, ht_away, ht_draw,
                                                   home_on, home_off, home_da, away_on, away_off, away_da, ht_score])
                    tr_ind += 2
                    first_blank_row += 1

                    if insert is True:
                        logging.info(f'--> <{game}> data is fetched and stored to gamedb!')
                        time.sleep(5)
                        save_to_csv()
                    else:
                        logging.info(f'--> <{game}> data is already exist in gamedb!')
                        save_to_csv()

        except Exception as ex:
            # print(ex)
            continue


def scanner2():
    # define 'driver' variable
    if check_excel_file() is False:
        return

    first_blank_row = get_first_blank_row()
    game = str(random.randint(1, 10))

    # iterate to extract game data
    while True:

        # insert data to gamedb
        insert = insert_or_update_row3(first_blank_row,
                                      ['current_date', game, 'home', '', '', 'ht_home', 'ht_away', 'ht_draw',
                                       'home_on', 'home_off', 'home_da', 'away_on', 'away_off', 'away_da', 'ht_score'])
        first_blank_row += 1
        print(insert)

        # if insert is True:
        #     logging.info(f'--> <{game}> data is fetched and stored to gamedb!')
        #     time.sleep(5)
        #     save_to_csv()
        # else:
        #     logging.info(f'--> <{game}> data is already exist in gamedb!')
        #     save_to_csv()

        time.sleep(5)


if __name__ == '__main__':
    logging.info('----------------- Script start running ... -----------------')
    scanner2()
    # v = get_first_blank_row3()
    # print(v)
