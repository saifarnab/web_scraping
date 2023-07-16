import subprocess
import pandas as pd
import logging
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import time
from datetime import date
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# subprocess.check_call(['pip', 'install', 'selenium'])
# subprocess.check_call(['pip', 'install', 'openpyxl'])

# excel db path
GAMEDB = 'gamedb.xlsx'

# log format
logging.basicConfig(
    format='%(asctime)s %(levelname)s %(message)s',
    datefmt='%H:%M:%S',
    level=logging.INFO)


def config_driver(maximize_window: bool) -> webdriver.Chrome:
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("lang=en-GB")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--proxy-server='direct://'")
    chrome_options.add_argument("--proxy-bypass-list=*")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=chrome_options)
    if maximize_window is True:
        driver.maximize_window()
    return driver


def create_excel_file():
    if os.path.exists(GAMEDB):
        return

    workbook = Workbook()
    sheet = workbook.active
    # Set column headers
    headers = ["GameDate", "Game", "Home", "HTHome", "HTAway", "HTDraw", "HomeOn", "HomeOff", "HomeDA", "AwayOn",
               "AwayOff", "AwayDA", "HTScore"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].font = Font(bold=True)

    workbook.save(GAMEDB)


def insert_or_update_row(new_data: list):
    df = pd.read_excel(GAMEDB)
    condition1 = df['GameDate'] == new_data[0]
    condition2 = df['Game'] == new_data[1]
    matching_rows = df.loc[condition1 & condition2]
    row_indexes = matching_rows.index

    if len(row_indexes) > 0:
        update_index = row_indexes[0]
        df.loc[update_index, 'GameDate'] = new_data[0]
        df.loc[update_index, 'Game'] = new_data[1]
        df.loc[update_index, 'Home'] = new_data[2]
        df.loc[update_index, 'HTHome'] = new_data[3]
        df.loc[update_index, 'HTAway'] = new_data[4]
        df.loc[update_index, 'HTDraw'] = new_data[5]
        df.loc[update_index, 'HomeOn'] = new_data[6]
        df.loc[update_index, 'HomeOff'] = new_data[7]
        df.loc[update_index, 'HomeDA'] = new_data[8]
        df.loc[update_index, 'AwayOn'] = new_data[9]
        df.loc[update_index, 'AwayOff'] = new_data[10]
        df.loc[update_index, 'AwayDA'] = new_data[11]
        df.loc[update_index, 'HTScore'] = new_data[12]
        df.to_excel(GAMEDB, index=False)


    else:
        new_row = {
            'GameDate': new_data[0],
            'Game': new_data[1],
            'Home': new_data[2],
            'HTHome': new_data[3],
            'HTAway': new_data[4],
            'HTDraw': new_data[5],
            'HomeOn': new_data[6],
            'HomeOff': new_data[7],
            'HomeDA': new_data[8],
            'AwayOn': new_data[9],
            'AwayOff': new_data[10],
            'AwayDA': new_data[11],
            'HTScore': new_data[12]
        }
        new_row_df = pd.DataFrame([new_row])
        df = pd.concat([df, new_row_df], ignore_index=True)
        df.to_excel(GAMEDB, index=False)


def scanner():
    # define 'driver' variable
    create_excel_file()
    driver = config_driver(True)

    # iterate to extract game data
    while True:

        try:
            driver.get('https://live.goalprofits.com/')
            # print(f'{pathlib.Path(__file__).parent.resolve()}\gamedb.gamedb;')

            if 'login' in driver.current_url:
                # login to the site
                username_input_button = driver.find_element(By.XPATH, '//input[@id="wlm_form_field_log"]')
                username_input_button.clear()
                username_input_button.send_keys("dennis.peibst@googlemail.com")
                pass_input_button = driver.find_element(By.XPATH, '//input[@id="wlm_form_field_pwd"]')
                pass_input_button.clear()
                pass_input_button.send_keys("Freelancer1")
                driver.find_element(By.XPATH, '//input[@id="wlm_form_field_wp-submit"]').click()
                logging.info("login success")
                time.sleep(2)
                # get to the live game section
                driver.get('https://live.goalprofits.com/')

            elif 'live' not in driver.current_url:
                logging.info('login or live not found')
                continue

            live_games_button = driver.find_element(By.XPATH, '//a[@href="#live"]')
            driver.execute_script("arguments[0].click();", live_games_button)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", live_games_button)
            time.sleep(1)

            print('ok 1')

            # find available live games
            table_elements = driver.find_elements(By.XPATH, '//div[@id="html2"]//table')
            print('total_table_elements: ', len(table_elements))

            # wait till 60s if no live games are available
            if len(table_elements) == 0:
                time.sleep(0)
                logging.info('--> no live games are playing, waiting 60s for next try.')
                continue

            # iterate each tr
            for table_ind in range(len(table_elements)):
                tr_elements = table_elements[table_ind].find_elements(By.XPATH, './/tbody//tr')
                tr_ind = 0

                # iterate each td
                for ind in range(int(len(tr_elements) / 2)):
                    td_elements = tr_elements[tr_ind].find_elements(By.XPATH, ".//td")
                    print('total_td_elements: ', len(td_elements))
                    current_date = date.today().strftime('%d/%m/%y')
                    try:
                        game_time = td_elements[0].text.strip()
                        logging.info(f'game_time --> {game_time}')
                    except Exception as exx:
                        game_time = 'N/A'

                    # if game time is not HT then return
                    # if game_time != 'HT':
                    #     tr_ind += 2
                    #     logging.info('Waiting to detect HT')
                    #     continue

                    # extract required data from 1st tr
                    team1 = td_elements[1].text.index('(')
                    team1 = td_elements[1].text[:team1].strip()
                    logging.info(f'team1 --> {team1}')

                    try:
                        ht_score = td_elements[2].text.split('\n')[0]
                        logging.info(f'ht_score --> {ht_score}')
                    except Exception as exx:
                        ht_score = 'N/A'

                    try:
                        v = td_elements[19].find_element(By.XPATH, './/a/img[@data-toggle="tooltip"]')
                        tooltip = v.get_attribute('data-original-title').replace('\n', '').replace(
                            "<div class='tooltip-prices'>", '').replace('<br>', '').replace('</div>', '').replace(
                            '<div>', '')
                        tooltip = "".join(tooltip.split())
                        ht_home, ht_draw, ht_away = tooltip[2:6], tooltip[8:12], tooltip[14:18]
                        # print(ht_home, ht_draw, ht_away)
                    except Exception as e:
                        ht_home, ht_draw, ht_away = 'N/A', 'N/A', 'N/A'

                    try:
                        home_on = td_elements[3].text
                        logging.info(f'home_on --> {home_on}')
                    except Exception as exx:
                        home_on = 'N/A'

                    try:
                        home_off = td_elements[4].text
                        logging.info(f'home_off --> {home_off}')
                    except Exception as exx:
                        home_off = 'N/A'

                    try:
                        home_da = td_elements[9].text
                        logging.info(f'home_da --> {home_da}')
                    except Exception as exx:
                        home_da = 'N/A'

                    # logging.info('extracted required data from 1st tr')

                    # extract required data from 2nd tr
                    td_elements2 = tr_elements[tr_ind + 1].find_elements(By.XPATH, ".//td")

                    team2 = td_elements2[0].text.index('(')
                    team2 = td_elements2[0].text[:team2].strip()
                    # logging.info(f'team2 --> {team2}')

                    try:
                        away_on = td_elements2[1].text
                        logging.info(f'away_on --> {away_on}')
                    except Exception as exx:
                        away_on = 'N/A'

                    try:
                        away_off = td_elements2[2].text
                        logging.info(f'away_off --> {away_off}')
                    except Exception as exx:
                        away_off = 'N/A'

                    try:
                        away_da = td_elements2[7].text
                        logging.info(f'away_da --> {away_da}')
                    except Exception as exx:
                        away_da = 'N/A'

                    # logging.info('extracted required data from 2nd tr')

                    game = team1 + ' v ' + team2

                    # extract data from i button which will open a modal
                    # td_elements[2].find_element(By.XPATH, './/a[2]').click()
                    driver.find_element(By.XPATH,
                                        "//img[@src='https://live.goalprofits.com/_layouts/images/info.png']").click()

                    # logging.info(f'i button 1st click done')
                    time.sleep(1)
                    driver.find_element(By.XPATH, '//a[@id="pregame-tab"]').click()
                    # logging.info(f'i button 2nd click done')
                    time.sleep(1)
                    try:
                        home = driver.find_elements(By.XPATH,
                                                    '//table[@class="table borderless table-striped score-table"]//span[@class="d-block font13 text-nowrap"]')[
                            0].text.split(':')[1].strip()
                        logging.info(f'home --> {home}')
                    except Exception as exx:
                        home = 'N/A'

                    # close the modal opened by i button
                    close_buttons = driver.find_elements(By.XPATH, '//button[@class="close"]')
                    for close_button in close_buttons:
                        try:
                            close_button.click()
                            # logging.info('clieck close button to close the modal')
                        except Exception as e:
                            continue

                    time.sleep(1)

                    # insert data to gamedb
                    insert_or_update_row([current_date, game, home, ht_home, ht_away, ht_draw, home_on,
                                          home_off, home_da, away_on, away_off, away_da, ht_score])
                    tr_ind += 2
                    logging.info(f'--> <{game}> data is fetched and stored to gamedb!')

        except Exception as ex:
            # print(ex)
            continue


if __name__ == '__main__':
    logging.info('----------------- Script start running ... -----------------')
    scanner()
